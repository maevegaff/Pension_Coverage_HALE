"""
05_additional_diagnostics.py

Four additional diagnostic checks not included in 02_analyze_data.py:

  1. VIF (Variance Inflation Factor)
  2. Breusch-Pagan test 
  3. GDP subsample table _> coefficient on log GDP per capita by income group,
     checking whether it moves in the expected direction (positive, larger in
     lower income groups)
  4. First-difference model — an alternative to fixed effects

Runs for BOTH pension models (contributory and social) and 
saves results to:
    outputs/additional_diagnostics_contributory.xlsx
    outputs/additional_diagnostics_social.xlsx

INPUT:  cleaned_data/cleaned_panel.csv  (produced by 01_clean_data.py)
Run:    python 05_additional_diagnostics.py
"""

import pandas as pd
import numpy as np
from linearmodels.panel import PanelOLS, FirstDifferenceOLS
from statsmodels.stats.outliers_influence import variance_inflation_factor
from scipy import stats
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
pd.set_option("display.width", 120)


# CONFIG


CLEANED_PANEL_PATH = "cleaned_data/cleaned_panel.csv"
OUT_DIR = "outputs"

DEPENDENT_VAR  = "hale_60"
CONTROL_VARS   = ["log_gdp_pc", "dependency_ratio", "oop_health_exp",
                   "gov_health_exp", "female_share_elderly", "urbanization_rate"]
ENTITY_COL     = "iso3"
TIME_COL       = "year"

INCOME_GROUPS  = ["Low income", "Lower middle income",
                   "Upper middle income", "High income"]

PENSION_MODELS = {
    "contributory": {
        "c_col":  "coverage_contrib_c",
        "sq_col": "coverage_contrib_c_sq",
        "label":  "Contributory Pension Coverage",
    },
    "social": {
        "c_col":  "coverage_social_c",
        "sq_col": "coverage_social_c_sq",
        "label":  "Social (Non-Contributory) Pension Coverage",
    },
}

VIF_FLAG_THRESHOLD = 10



# HELPERS


def load_panel():
    df = pd.read_csv(CLEANED_PANEL_PATH)
    df = df.set_index([ENTITY_COL, TIME_COL])
    return df


def stars(p):
    if pd.isna(p) or p is None: return ""
    if p < 0.01: return "***"
    if p < 0.05: return "**"
    if p < 0.10: return "*"
    return ""


def get_model_df(df, model_spec):
    """Returns the modeling sample for a given pension model."""
    c_col  = model_spec["c_col"]
    sq_col = model_spec["sq_col"]
    income_dummies = [c for c in df.columns
                       if c.startswith("income_") and c != "income_group"]
    interactions   = [c for c in df.columns
                       if c.startswith(c_col.replace("_c", "_x_income_"))]
    regressors = [c_col, sq_col] + CONTROL_VARS + income_dummies + interactions
    return df.dropna(subset=[DEPENDENT_VAR] + regressors), regressors



# DIAGNOSTIC 1: VIF


def run_vif(df, model_spec):
    """
    VIF is computed on the demeaned (within-transformed) regressor matrix,
    since that is what the FE estimator actually works with. Values above 10
    indicate potentially problematic multicollinearity.
    """
    print(f"\n[Diag 1 -- VIF] {model_spec['label']}")
    model_df, regressors = get_model_df(df, model_spec)

    # Within-demean by entity (approximates what FE does)
    X = model_df[regressors].copy()
    X_demeaned = X - X.groupby(level=ENTITY_COL).transform("mean")
    X_demeaned = X_demeaned.dropna()

    # Drop any columns with zero variance after demeaning (absorbed by FE)
    zero_var = X_demeaned.std() == 0
    X_demeaned = X_demeaned.loc[:, ~zero_var]

    results = []
    cols = list(X_demeaned.columns)
    X_arr = X_demeaned.values

    for i, col in enumerate(cols):
        try:
            vif = variance_inflation_factor(X_arr, i)
        except Exception:
            vif = np.nan
        flag = "⚠ FLAG" if (not np.isnan(vif) and vif > VIF_FLAG_THRESHOLD) else ""
        results.append({"Variable": col, "VIF": round(vif, 2), "Flag": flag})
        print(f"  {col}: VIF = {vif:.2f} {flag}")

    flagged = [r for r in results if r["Flag"]]
    if flagged:
        print(f"\n  ⚠ {len(flagged)} variable(s) exceed VIF threshold of "
              f"{VIF_FLAG_THRESHOLD}: {[r['Variable'] for r in flagged]}")
    else:
        print(f"\n  All VIFs below {VIF_FLAG_THRESHOLD} — no severe multicollinearity detected")

    return results



# DIAGNOSTIC 2: BREUSCH-PAGAN TEST


def run_breusch_pagan(df, model_spec):
    """
    Tests for heteroskedasticity in the FE residuals. A significant p-value
    (p < 0.05) means the error variance is not constant -- robust (clustered)
    standard errors are then essential, which our primary model already uses.
    """
    print(f"\n[Diag 2 -- Breusch-Pagan] {model_spec['label']}")
    model_df, regressors = get_model_df(df, model_spec)

    y = model_df[DEPENDENT_VAR]
    X = model_df[regressors]

    res = PanelOLS(y, X, entity_effects=True, time_effects=True,
                    drop_absorbed=True).fit(cov_type="unadjusted")

    residuals = res.resids.values
    fitted    = res.fitted_values.values

    # Breusch-Pagan: regress squared residuals on fitted values
    resid_sq = residuals ** 2
    fitted_c = np.column_stack([np.ones(len(fitted)), fitted])
    bp_res   = np.linalg.lstsq(fitted_c, resid_sq, rcond=None)
    bp_coefs = bp_res[0]

    n      = len(residuals)
    r2_aux = 1 - np.sum((resid_sq - fitted_c @ bp_coefs) ** 2) / \
             np.sum((resid_sq - resid_sq.mean()) ** 2)
    bp_stat = n * r2_aux
    bp_p    = 1 - stats.chi2.cdf(bp_stat, df=1)

    result = {
        "Test":           "Breusch-Pagan",
        "Test Statistic": round(bp_stat, 4),
        "Degrees of Freedom": 1,
        "p-value":        round(bp_p, 4),
        "Significance":   stars(bp_p),
        "Interpretation": (
            "Heteroskedasticity detected — clustered SEs (already used in primary model) are appropriate"
            if bp_p < 0.05 else
            "No significant heteroskedasticity detected at 5% level"
        ),
    }
    print(f"  BP statistic = {bp_stat:.4f}, p = {bp_p:.4f}")
    print(f"  -> {result['Interpretation']}")
    return result



# DIAGNOSTIC 3: GDP SUBSAMPLE TABLE BY INCOME GROUP


def run_gdp_subsample(df, model_spec):
    """
    Runs the primary FE model separately for each income group and extracts
    the log GDP per capita coefficient. Expected direction: positive in all
    groups (richer countries have better HALE), with the effect potentially
    larger in lower income groups where the marginal return to wealth is
    higher (consistent with Deaton 2003 / Wagstaff & van Doorslaer 2000).
    """
    print(f"\n[Diag 3 -- GDP Subsample] {model_spec['label']}")
    c_col  = model_spec["c_col"]
    sq_col = model_spec["sq_col"]
    split_regressors = [c_col, sq_col] + CONTROL_VARS
    results = []

    for group in INCOME_GROUPS:
        sub = df[df["income_group"] == group].dropna(
            subset=[DEPENDENT_VAR] + split_regressors)
        n_countries = sub.index.get_level_values(ENTITY_COL).nunique()

        row = {
            "Income Group":    group,
            "GDP Coefficient": None,
            "Std. Error":      None,
            "p-value":         None,
            "Sig.":            "",
            "N (obs)":         len(sub),
            "N (countries)":   n_countries,
            "Direction":       "",
            "Note":            "",
        }

        if n_countries < 3:
            row["Note"] = f"Too few countries ({n_countries})"
            print(f"  {group}: too few countries, skipping")
            results.append(row)
            continue

        try:
            y, X = sub[DEPENDENT_VAR], sub[split_regressors]
            res  = PanelOLS(y, X, entity_effects=True, time_effects=True,
                             drop_absorbed=True).fit(cov_type="clustered",
                                                      cluster_entity=True)
            coef = res.params["log_gdp_pc"]
            se   = res.std_errors["log_gdp_pc"]
            p    = res.pvalues["log_gdp_pc"]
            row.update({
                "GDP Coefficient": round(coef, 4),
                "Std. Error":      round(se, 4),
                "p-value":         round(p, 4),
                "Sig.":            stars(p),
                "Direction":       "✓ Positive (expected)" if coef > 0
                                   else "✗ Negative (unexpected)",
            })
            print(f"  {group}: GDP coef = {coef:.4f}, p = {p:.4f} "
                  f"({'positive ✓' if coef > 0 else 'negative ✗'})")
        except Exception as e:
            row["Note"] = f"Model failed: {e}"
            print(f"  {group}: failed ({e})")

        results.append(row)
    return results



# DIAGNOSTIC 4: FIRST-DIFFERENCE MODEL

def run_first_difference(df, model_spec):
    """
    First-difference (FD) model: differences out country fixed effects by
    subtracting year t-1 from year t within each country. Consistent with
    FE under strict exogeneity but more efficient when errors follow a
    random walk. 
   
    """
    print(f"\n[Diag 4 -- First Difference] {model_spec['label']}")
    c_col  = model_spec["c_col"]
    sq_col = model_spec["sq_col"]
    regressors = [c_col, sq_col] + CONTROL_VARS

    model_df = df.dropna(subset=[DEPENDENT_VAR] + regressors).copy()
    model_df = model_df.sort_index()

    # Manual first-differencing
    diff_cols = [DEPENDENT_VAR] + regressors
    diff = model_df[diff_cols].groupby(level=ENTITY_COL).diff()
    diff = diff.dropna()

    # Add year dummies to proxy for common time shocks
    year_idx = diff.index.get_level_values(TIME_COL)
    year_dummies = pd.get_dummies(year_idx, prefix="yr", drop_first=True).astype(float)
    year_dummies.index = diff.index

    X_fd = pd.concat([diff[regressors], year_dummies], axis=1)
    y_fd = diff[DEPENDENT_VAR]
    valid = X_fd.notna().all(axis=1) & y_fd.notna()
    X_fd, y_fd = X_fd[valid], y_fd[valid]

    n_obs      = len(y_fd)
    n_countries = y_fd.index.get_level_values(ENTITY_COL).nunique()

    result = {
        "Coefficient":     None, "Std. Error":    None,
        "p-value":         None, "Sig.":          "",
        "R² (FD model)":   None, "N (obs)":       n_obs,
        "N (countries)":   n_countries, "Note":   "",
        "Consistent with FE?": "",
    }

    try:
        from sklearn.linear_model import LinearRegression
        # OLS on differenced data with clustered SEs approximated via HC3
        X_arr = X_fd.values
        y_arr = y_fd.values
        Xt    = X_arr.T
        beta  = np.linalg.lstsq(X_arr, y_arr, rcond=None)[0]
        yhat  = X_arr @ beta
        resid = y_arr - yhat

        # HC3 sandwich for approximate clustered SEs
        n, k  = X_arr.shape
        XtX_inv = np.linalg.pinv(Xt @ X_arr)
        meat  = sum(
            (X_arr[i:i+1].T @ X_arr[i:i+1]) * resid[i] ** 2
            for i in range(n))
        sandwich = XtX_inv @ meat @ XtX_inv
        se_all  = np.sqrt(np.diag(sandwich))

        coef = beta[0]  # pension coverage is first regressor
        se   = se_all[0]
        t    = coef / se if se > 0 else np.nan
        p    = 2 * (1 - stats.t.cdf(abs(t), df=n - k)) if not np.isnan(t) else np.nan

        ss_res = np.sum(resid ** 2)
        ss_tot = np.sum((y_arr - y_arr.mean()) ** 2)
        r2     = 1 - ss_res / ss_tot if ss_tot > 0 else np.nan

        result.update({
            "Coefficient":   round(coef, 4),
            "Std. Error":    round(se, 4),
            "p-value":       round(p, 4) if not np.isnan(p) else None,
            "Sig.":          stars(p),
            "R² (FD model)": round(r2, 4),
        })

        print(f"  FD coef = {coef:.4f}, SE = {se:.4f}, p = {p:.4f}")

    except Exception as e:
        result["Note"] = f"Model failed: {e}"
        print(f"  FD model failed: {e}")

    return result



# EXCEL OUTPUT


def write_diagnostics_excel(vif_res, bp_res, gdp_res, fd_res, label, out_path):
    wb = Workbook()
    header_fill = PatternFill("solid", start_color="1F4E5F", end_color="1F4E5F")
    flag_fill   = PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0")
    warn_fill   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    alt_fill    = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    hf  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    nf  = Font(name="Arial", size=10)
    bf  = Font(name="Arial", bold=True, size=10)
    wf  = Font(name="Arial", italic=True, size=9, color="7F6000")
    thin = Side(style="thin")

    def header_row(ws, row_num, headers, col_widths=None):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=c, value=h)
            cell.font = hf; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin)
        if col_widths:
            for c, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(c)].width = w

    def title_row(ws, text, n_cols):
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        ws["A1"].value = text
        ws["A1"].font  = Font(name="Arial", bold=True, size=11)
        ws["A1"].alignment = Alignment(horizontal="center")

    # ---- Sheet 1: VIF ----
    ws1 = wb.active
    ws1.title = "Diag 1 VIF"
    title_row(ws1, f"{label} — VIF (Variance Inflation Factors, entity-demeaned regressors)", 3)
    ws1.merge_cells("A2:C2")
    ws1["A2"].value = (f"Values above {VIF_FLAG_THRESHOLD} flagged in red. "
                        "Computed on within-transformed (demeaned) regressors "
                        "to match what the FE estimator uses.")
    ws1["A2"].font = Font(name="Arial", italic=True, size=9)
    ws1["A2"].alignment = Alignment(horizontal="center")
    header_row(ws1, 3, ["Variable", "VIF", "Flag"], [32, 10, 12])
    for i, row in enumerate(vif_res, start=4):
        is_flag = row["Flag"] != ""
        fill = flag_fill if is_flag else (alt_fill if i % 2 == 0 else None)
        for c, key in enumerate(["Variable", "VIF", "Flag"], start=1):
            cell = ws1.cell(row=i, column=c, value=row[key])
            cell.font = Font(name="Arial", bold=is_flag, size=10,
                              color="C00000" if is_flag else "000000")
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center")
            if fill: cell.fill = fill
    ws1.freeze_panes = "A4"

    #  Sheet 2: Breusch-Pagan 
    ws2 = wb.create_sheet("Diag 2 Breusch-Pagan")
    title_row(ws2, f"{label} — Breusch-Pagan Test for Heteroskedasticity", 2)
    ws2.merge_cells("A2:B2")
    ws2["A2"].value = ("H0: constant error variance (homoskedasticity). "
                        "Significant result (p<0.05) supports using clustered SEs.")
    ws2["A2"].font = Font(name="Arial", italic=True, size=9)
    ws2["A2"].alignment = Alignment(horizontal="center")
    bp_rows = [
        ("Test", bp_res["Test"]),
        ("Test Statistic (LM)", bp_res["Test Statistic"]),
        ("Degrees of Freedom", bp_res["Degrees of Freedom"]),
        ("p-value", bp_res["p-value"]),
        ("Significance", bp_res["Significance"]),
        ("Interpretation", bp_res["Interpretation"]),
    ]
    header_row(ws2, 3, ["Statistic", "Value"], [28, 55])
    for i, (stat, val) in enumerate(bp_rows, start=4):
        ws2.cell(row=i, column=1, value=stat).font = bf
        cell = ws2.cell(row=i, column=2, value=val)
        cell.font = nf
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
        if i % 2 == 0:
            ws2.cell(row=i, column=1).fill = alt_fill
            cell.fill = alt_fill
    ws2.row_dimensions[9].height = 30
    ws2.freeze_panes = "A4"

    #  Sheet 3: GDP Subsample 
    ws3 = wb.create_sheet("Diag 3 GDP Subsample")
    title_row(ws3, f"{label} — Log GDP per Capita Coefficient by Income Group", 8)
    ws3.merge_cells("A2:H2")
    ws3["A2"].value = ("H0: GDP coefficient is positive in all income groups. "
                        "Expected: larger in lower income groups (diminishing returns, "
                        "per Deaton 2003). Two-way FE, clustered SEs.")
    ws3["A2"].font = Font(name="Arial", italic=True, size=9)
    ws3["A2"].alignment = Alignment(horizontal="center")
    header_row(ws3, 3,
               ["Income Group", "GDP Coefficient", "Std. Error", "p-value",
                "Sig.", "N (obs)", "N (countries)", "Direction"],
               [24, 16, 12, 10, 6, 10, 14, 24])
    for i, row in enumerate(gdp_res, start=4):
        is_warn = bool(row.get("Note"))
        is_unexpected = str(row.get("Direction", "")).startswith("✗")
        fill = (warn_fill if is_warn else
                flag_fill if is_unexpected else
                alt_fill if i % 2 == 0 else None)
        keys = ["Income Group","GDP Coefficient","Std. Error","p-value",
                "Sig.","N (obs)","N (countries)","Direction"]
        for c, key in enumerate(keys, start=1):
            val  = row.get("Note") if (is_warn and c == 8) else row.get(key, "")
            cell = ws3.cell(row=i, column=c, value=val)
            cell.font = wf if is_warn else nf
            cell.alignment = Alignment(horizontal="left" if c in [1,8] else "center")
            if fill: cell.fill = fill
    ws3.freeze_panes = "A4"

    # Sheet 4: First Difference 
    ws4 = wb.create_sheet("Diag 4 First Difference")
    title_row(ws4, f"{label} — First-Difference Model (Alternative to FE)", 2)
    ws4.merge_cells("A2:B2")
    ws4["A2"].value = ("FD removes country fixed effects by differencing. "
                        "Coefficient should be directionally consistent with "
                        "primary FE result. Year dummies included to proxy for "
                        "common time shocks.")
    ws4["A2"].font = Font(name="Arial", italic=True, size=9)
    ws4["A2"].alignment = Alignment(horizontal="center")
    fd_rows = [
        ("FD Coefficient (pension coverage)", fd_res.get("Coefficient")),
        ("Std. Error",                        fd_res.get("Std. Error")),
        ("p-value",                           fd_res.get("p-value")),
        ("Significance",                      fd_res.get("Sig.")),
        ("R² (FD model)",                     fd_res.get("R² (FD model)")),
        ("N (observations)",                  fd_res.get("N (obs)")),
        ("N (countries)",                     fd_res.get("N (countries)")),
        ("Note",                              fd_res.get("Note", "")),
    ]
    header_row(ws4, 3, ["Statistic", "Value"], [36, 24])
    for i, (stat, val) in enumerate(fd_rows, start=4):
        ws4.cell(row=i, column=1, value=stat).font = bf
        cell = ws4.cell(row=i, column=2, value=val)
        cell.font = nf
        cell.alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            ws4.cell(row=i, column=1).fill = alt_fill
            cell.fill = alt_fill
    ws4.freeze_panes = "A4"

    wb.save(out_path)
    print(f"\nSaved to {out_path}")



# MAIN


def main():
    import os
    os.makedirs(OUT_DIR, exist_ok=True)

    df = load_panel()

    for model_key, model_spec in PENSION_MODELS.items():
        label = model_spec["label"]
        print("\n" + "#" * 65)
        print(f"# {label.upper()}")
        print("#" * 65)

        vif_res = run_vif(df, model_spec)
        bp_res  = run_breusch_pagan(df, model_spec)
        gdp_res = run_gdp_subsample(df, model_spec)
        fd_res  = run_first_difference(df, model_spec)

        out_path = f"{OUT_DIR}/additional_diagnostics_{model_key}.xlsx"
        write_diagnostics_excel(vif_res, bp_res, gdp_res, fd_res, label, out_path)

    print("\nDone. Output files:")
    for m in PENSION_MODELS:
        print(f"  outputs/additional_diagnostics_{m}.xlsx")


if __name__ == "__main__":
    main()
