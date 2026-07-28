"""
02_analyze_data.py

 Descriptives -> Hausman test -> primary two-way FE model
-> interpretation -> robustness checks (lags, split-sample, IV).

INPUT:  cleaned_data/cleaned_panel.csv  (produced by 01_clean_data.py)
OUTPUT: Printed results + result summary CSVs in outputs/ (one set per
        pension variable, e.g. step10_primary_model_summary_contributory.csv)

Run: python 02_analyze_data.py
"""

import pandas as pd
import numpy as np
from scipy import stats
from linearmodels.panel import PanelOLS, RandomEffects
from linearmodels.iv import IV2SLS
import warnings

warnings.filterwarnings("ignore", category=FutureWarning)
pd.set_option("display.width", 120)


# CONFIG


CLEANED_PANEL_PATH = r"C:\Users\maeve\Downloads\FinDiss\data\cleaned_panel.csv"
OUT_DIR = r"C:\Users\maeve\Downloads\FinDiss\data\outputs"

DEPENDENT_VAR = "hale_60"
CONTROL_VARS = ["log_gdp_pc", "dependency_ratio", "oop_health_exp",
                 "gov_health_exp", "female_share_elderly", "urbanization_rate"]
INCOME_DUMMY_PREFIX = "income_"
INSTRUMENT_VAR = "dependency_ratio"
LAG_YEARS = [1, 2, 3]

ENTITY_COL = "iso3"
TIME_COL = "year"

# The two pension models to run, each fully independent of the other.
PENSION_MODELS = {
    "contributory": {
        "raw_col": "pension_contributory",
        "c_col": "coverage_contrib_c",
        "sq_col": "coverage_contrib_c_sq",
        "interaction_prefix": "coverage_contrib_x_income_",
        "label": "Contributory pension coverage",
    },
    "social": {
        "raw_col": "pension_social",
        "c_col": "coverage_social_c",
        "sq_col": "coverage_social_c_sq",
        "interaction_prefix": "coverage_social_x_income_",
        "label": "Social (non-contributory) pension coverage",
    },
}



# HELPERS

def load_panel():
    df = pd.read_csv(CLEANED_PANEL_PATH)
    df = df.set_index([ENTITY_COL, TIME_COL])
    return df


def get_income_cols(df, interaction_prefix):
    income_dummies = [c for c in df.columns
                       if c.startswith(INCOME_DUMMY_PREFIX) and c != "income_group"]
    interactions = [c for c in df.columns if c.startswith(interaction_prefix)]
    return income_dummies, interactions


def build_regressors(df, model_spec):
    income_dummies, interactions = get_income_cols(df, model_spec["interaction_prefix"])
    return [model_spec["c_col"], model_spec["sq_col"]] + CONTROL_VARS + income_dummies + interactions



# DESCRIPTIVE STATISTICS (both pension variables shown together for context)


def descriptive_checks(df):
    print("=" * 70)
    print("[Step 8] Descriptive statistics (both pension variables, for context)")
    print("=" * 70)

    key_cols = [DEPENDENT_VAR, "pension_contributory", "pension_social"] + CONTROL_VARS
    desc = df[key_cols].describe().T[["count", "mean", "std", "min", "max"]]
    print(desc.round(2))

    print("\nCorrelation between contributory and social coverage (where both present):")
    both = df[["pension_contributory", "pension_social"]].dropna()
    if len(both) > 2:
        print(f"  r = {both.corr().iloc[0, 1]:.3f} (n={len(both)})")
    else:
        print(f"  Too few overlapping rows (n={len(both)}) to compute meaningfully")

    desc.to_csv(f"{OUT_DIR}/step8_descriptives.csv")
    return desc



# HAUSMAN TEST

def hausman_test(df, regressors, label):
    print(f"\n[Step 9 -- {label}] Hausman test (fixed effects vs random effects)")

    y = df[DEPENDENT_VAR]
    X = df[regressors]

    fe_res = PanelOLS(y, X, entity_effects=True, time_effects=True,
                       drop_absorbed=True).fit(cov_type="clustered", cluster_entity=True)
    re_res = RandomEffects(y, X).fit(cov_type="clustered", cluster_entity=True)

    common_params = [p for p in fe_res.params.index if p in re_res.params.index]
    b_diff = fe_res.params[common_params] - re_res.params[common_params]
    v_diff = (fe_res.cov.loc[common_params, common_params] -
              re_res.cov.loc[common_params, common_params])

    try:
        v_diff_inv = np.linalg.inv(v_diff.values)
        h_stat = float(b_diff.values @ v_diff_inv @ b_diff.values.T)
        df_h = len(common_params)
        p_value = 1 - stats.chi2.cdf(h_stat, df_h)
        print(f"  Hausman statistic: {h_stat:.3f} (df={df_h}) p-value: {p_value:.4f}")
        if p_value < 0.05:
            print("  -> p < 0.05: fixed effects preferred")
        else:
            print("  -> p >= 0.05: cannot reject RE at conventional levels; "
                  "FE retained per literature-based justification")
    except np.linalg.LinAlgError:
        print("  -> Covariance difference not invertible (common with small "
              "samples / near-collinear regressors). Report FE vs RE "
              "coefficients qualitatively instead.")

    return fe_res, re_res



# PRIMARY TWO-WAY FIXED EFFECTS MODEL 

def primary_model(df, regressors, label):
    print(f"\n[Step 10 -- {label}] Primary model: two-way fixed effects, clustered SEs")
    y = df[DEPENDENT_VAR]
    X = df[regressors]
    res = PanelOLS(y, X, entity_effects=True, time_effects=True,
                    drop_absorbed=True).fit(cov_type="clustered", cluster_entity=True)
    print(res.summary)
    return res



# INTERPRETATION


def interpret_results(res, model_spec, label):
    print(f"\n[Step 11 -- {label}] Interpretation against hypotheses")

    def report(varname, h_label):
        if varname not in res.params.index:
            print(f"  {h_label}: '{varname}' not in model (possibly absorbed)")
            return
        coef, pval = res.params[varname], res.pvalues[varname]
        sig = "significant" if pval < 0.05 else "not significant"
        print(f"  {h_label}: coef = {coef:.4f}, p = {pval:.4f} ({sig} at 5%)")

    print(f"\nH1 -- {model_spec['label']} effect on HALE:")
    report(model_spec["c_col"], model_spec["label"])

    print(f"\nH2 -- diminishing returns ({model_spec['label']}):")
    report(model_spec["sq_col"], f"{model_spec['label']} squared")

    print(f"\nH1 (moderation) -- income group interactions ({model_spec['label']}):")
    for v in res.params.index:
        if v.startswith(model_spec["interaction_prefix"]):
            report(v, f"  {v}")



# ROBUSTNESS CHECKS


def robustness_lagged(df, regressors, model_spec, label):
    print(f"\n[Step 12a -- {label}] Robustness: lagged coverage (1-3 years)")
    df = df.sort_index()
    c_col = model_spec["c_col"]

    for lag in LAG_YEARS:
        lagged_col = f"{c_col}_lag{lag}"
        df[lagged_col] = df.groupby(level=ENTITY_COL)[c_col].shift(lag)
        lag_regressors = [r if r != c_col else lagged_col for r in regressors]
        sub = df.dropna(subset=[DEPENDENT_VAR, lagged_col] + CONTROL_VARS)

        if sub.index.get_level_values(ENTITY_COL).nunique() < 5:
            print(f"  Lag {lag}: too few countries with data ({sub.index.get_level_values(ENTITY_COL).nunique()}), skipping")
            continue

        y, X = sub[DEPENDENT_VAR], sub[lag_regressors]
        try:
            res = PanelOLS(y, X, entity_effects=True, time_effects=True,
                            drop_absorbed=True).fit(cov_type="clustered", cluster_entity=True)
            print(f"  Lag {lag}: coef = {res.params[lagged_col]:.4f}, "
                  f"p = {res.pvalues[lagged_col]:.4f}, N = {res.nobs}")
        except Exception as e:
            print(f"  Lag {lag}: model failed ({e})")


def robustness_split_sample(df, model_spec, label):
    print(f"\n[Step 12b -- {label}] Robustness: split-sample by income group")
    c_col = model_spec["sq_col"].replace("_sq", "")  # back to c_col, just for clarity
    c_col = model_spec["c_col"]
    split_regressors = [model_spec["sq_col"], c_col] + CONTROL_VARS

    if "income_group" not in df.columns:
        print("  -> 'income_group' not found, skipping")
        return

    for group in df["income_group"].dropna().unique():
        sub = df[df["income_group"] == group].dropna(subset=[DEPENDENT_VAR] + split_regressors)
        n_countries = sub.index.get_level_values(ENTITY_COL).nunique()
        if n_countries < 3:
            print(f"  {group}: too few countries ({n_countries}), skipping")
            continue
        y, X = sub[DEPENDENT_VAR], sub[split_regressors]
        try:
            res = PanelOLS(y, X, entity_effects=True, time_effects=True,
                            drop_absorbed=True).fit(cov_type="clustered", cluster_entity=True)
            print(f"  {group}: coef = {res.params[c_col]:.4f}, "
                  f"p = {res.pvalues[c_col]:.4f}, N = {res.nobs}, countries = {n_countries}")
        except Exception as e:
            print(f"  {group}: model failed ({e})")


def robustness_iv(df, model_spec, label):
    print(f"\n[Step 12c -- {label}] Robustness: instrumental variable (2SLS)")
    c_col = model_spec["c_col"]
    print(f"  Endogenous regressor: {c_col}")
    print(f"  Instrument: {INSTRUMENT_VAR}")

    exog_controls = [c for c in CONTROL_VARS if c != INSTRUMENT_VAR]
    cols_needed = [DEPENDENT_VAR, c_col, INSTRUMENT_VAR] + exog_controls
    sub = df.dropna(subset=cols_needed).copy()

    n_countries = sub.index.get_level_values(ENTITY_COL).nunique()
    if n_countries < 10:
        print(f"  -> Only {n_countries} countries with complete data for this "
              f"check -- too few for a reliable two-way-demeaned IV estimate, skipping.")
        return

    def two_way_demean(frame, cols, n_iter=10):
        out = frame[cols].copy()
        for _ in range(n_iter):
            out = out - out.groupby(level=ENTITY_COL).transform("mean") + out.mean()
            out = out - out.groupby(level=TIME_COL).transform("mean") + out.mean()
        return out

    demeaned = two_way_demean(sub, cols_needed)

    y = demeaned[DEPENDENT_VAR]
    endog = demeaned[[c_col]]
    exog = demeaned[exog_controls] if exog_controls else None
    instr = demeaned[[INSTRUMENT_VAR]]

    try:
        iv_res = IV2SLS(dependent=y, exog=exog, endog=endog, instruments=instr).fit(
            cov_type="clustered", clusters=sub.index.get_level_values(ENTITY_COL))
        print(iv_res.summary)

        fs_y = demeaned[c_col]
        fs_X = pd.concat([demeaned[exog_controls], instr], axis=1) if exog_controls else instr
        fs_res = PanelOLS(fs_y, fs_X).fit(cov_type="clustered", cluster_entity=True)
        instr_t = fs_res.tstats[INSTRUMENT_VAR]
        print(f"\n  First-stage instrument check: t-stat = {instr_t:.2f}, "
              f"p = {fs_res.pvalues[INSTRUMENT_VAR]:.4f}")
        if abs(instr_t) < 3.16:
            print("  -> WARNING: instrument may be weak (rule-of-thumb F < 10).")
        else:
            print("  -> Instrument passes rule-of-thumb strength check (F > 10).")
    except Exception as e:
        print(f"  -> IV model failed to estimate ({e})")



# RUN ONE FULL MODEL FOR ONE PENSION VARIABLE

def run_pension_model(df, model_key, model_spec):
    label = model_spec["label"]
    print("\n" + "#" * 70)
    print(f"# MODEL: {label.upper()}")
    print("#" * 70)

    regressors = build_regressors(df, model_spec)
    required = [DEPENDENT_VAR] + regressors
    model_df = df.dropna(subset=required)
    n_countries = model_df.index.get_level_values(ENTITY_COL).nunique()
    print(f"\nModeling sample for {label}: {len(model_df):,} country-year "
          f"observations, {n_countries} countries")

    if n_countries < 10:
        print(f"  -> Only {n_countries} countries available -- too few to "
              f"estimate a two-way fixed effects model reliably. Skipping "
              f"this model. Consider pooled OLS or reporting descriptively instead.")
        return None

    hausman_test(model_df, regressors, label)
    primary_res = primary_model(model_df, regressors, label)
    interpret_results(primary_res, model_spec, label)
    robustness_lagged(model_df, regressors, model_spec, label)
    robustness_split_sample(model_df, model_spec, label)
    robustness_iv(model_df, model_spec, label)

    #  Save FULL coefficient table (all variables including controls) 
    # Human-readable variable labels for the results table
    VARIABLE_LABELS = {
        model_spec["c_col"]:           "Pension coverage (centered)",
        model_spec["sq_col"]:          "Pension coverage² (centered)",
        "log_gdp_pc":                  "GDP per capita (log, PPP)",
        "dependency_ratio":            "Old-age dependency ratio",
        "oop_health_exp":              "OOP health expenditure (% CHE)",
        "gov_health_exp":              "Gov't health expenditure (% GDP)",
        "female_share_elderly":        "Female share of elderly population (%)",
        "urbanization_rate":           "Urbanisation rate (%)",
    }
    for int_col in [c for c in model_df.columns if c.startswith(model_spec["interaction_prefix"])]:
        group = int_col.replace(model_spec["interaction_prefix"], "").replace("_", " ").title()
        VARIABLE_LABELS[int_col] = f"Coverage × {group}"

    # Significance stars
    def stars(p):
        if p < 0.01:  return "***"
        if p < 0.05:  return "**"
        if p < 0.10:  return "*"
        return ""

    # Ordered variable list: pension vars first, then controls, then interactions
    ordered_vars = (
        [model_spec["c_col"], model_spec["sq_col"]]
        + CONTROL_VARS
        + [c for c in model_df.columns if c.startswith(model_spec["interaction_prefix"])]
    )

    summary_rows = []
    for v in ordered_vars:
        if v not in primary_res.params.index:
            continue
        coef = primary_res.params[v]
        se   = primary_res.std_errors[v]
        p    = primary_res.pvalues[v]
        summary_rows.append({
            "Variable":      VARIABLE_LABELS.get(v, v),
            "Coefficient":   round(coef, 4),
            "Std. Error":    round(se, 4),
            "p-value":       round(p, 4),
            "Significance":  stars(p),
        })

    full_df = pd.DataFrame(summary_rows)

    # Append model-fit stats as footer rows
    r2     = round(primary_res.rsquared, 4)
    n_obs  = primary_res.nobs
    n_ctry = model_df.index.get_level_values(ENTITY_COL).nunique()
    for label_row, val in [("R² (within)", r2), ("N (observations)", n_obs),
                            ("N (countries)", n_ctry)]:
        summary_rows.append({"Variable": label_row, "Coefficient": val,
                              "Std. Error": "", "p-value": "", "Significance": ""})

    # Save CSV (machine-readable)
    csv_path = f"{OUT_DIR}/step10_primary_model_summary_{model_key}.csv"
    full_df.to_csv(csv_path, index=False)

    # Save formatted Excel table (for pasting into write-up)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"{model_key.title()} Model"

    # Title row
    title = f"Table: {label} — Primary Two-Way Fixed Effects Model (HALE at Age 60)"
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Arial", bold=True, size=11)
    title_cell.alignment = Alignment(horizontal="center")

    # Header row
    headers = ["Variable", "Coefficient", "Std. Error", "p-value", "Sig."]
    header_fill = PatternFill("solid", start_color="1F4E5F", end_color="1F4E5F")
    thin = Side(style="thin")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    # Data rows
    divider_vars = {model_spec["sq_col"]: True,
                    "urbanization_rate": True}
    alt_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")

    data_rows = [r for r in summary_rows if r["Variable"] not in
                 ["R² (within)", "N (observations)", "N (countries)"]]
    fit_rows  = [r for r in summary_rows if r["Variable"] in
                 ["R² (within)", "N (observations)", "N (countries)"]]

    for i, row in enumerate(data_rows, start=3):
        fill = alt_fill if i % 2 == 0 else None
        for c, key in enumerate(["Variable", "Coefficient", "Std. Error", "p-value", "Significance"], start=1):
            cell = ws.cell(row=i, column=c, value=row.get(key, ""))
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center")
            if fill:
                cell.fill = fill

    # Separator line before fit stats
    sep_row = 3 + len(data_rows)
    for c in range(1, 6):
        ws.cell(row=sep_row, column=c).border = Border(top=thin)

    # Fit stat rows
    for i, row in enumerate(fit_rows, start=sep_row):
        ws.cell(row=i, column=1, value=row["Variable"]).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=i, column=2, value=row["Coefficient"]).font = Font(name="Arial", size=10)
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")

    # Note row
    note_row = sep_row + len(fit_rows)
    ws.merge_cells(f"A{note_row}:E{note_row}")
    note = ("Notes: Two-way fixed effects (country + year FE), standard errors "
            "clustered at country level. *** p<0.01, ** p<0.05, * p<0.10. "
            "Coverage variable is mean-centered. Country and year fixed effects "
            "not reported.")
    note_cell = ws.cell(row=note_row, column=1, value=note)
    note_cell.font = Font(name="Arial", italic=True, size=9)
    note_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 40

    # Column widths
    col_widths = [42, 14, 12, 10, 8]
    for c, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A3"

    xlsx_path = f"{OUT_DIR}/results_table_{model_key}.xlsx"
    wb.save(xlsx_path)
    print(f"\nSaved full coefficient table to {csv_path}")
    print(f"Saved formatted results table to {xlsx_path}")

    return primary_res


# MAIN


def main():
    import os
    os.makedirs(OUT_DIR, exist_ok=True)

    df = load_panel()
    descriptive_checks(df)

    results = {}
    for model_key, model_spec in PENSION_MODELS.items():
        results[model_key] = run_pension_model(df, model_key, model_spec)

    print("\n" + "=" * 70)
    print("SUMMARY ACROSS BOTH MODELS")
    print("=" * 70)
    for model_key, res in results.items():
        label = PENSION_MODELS[model_key]["label"]
        if res is None:
            print(f"  {label}: model skipped (insufficient data)")
        else:
            c_col = PENSION_MODELS[model_key]["c_col"]
            coef, pval = res.params[c_col], res.pvalues[c_col]
            print(f"  {label}: coef = {coef:.4f}, p = {pval:.4f}, N = {res.nobs}")


if __name__ == "__main__":
    main()
