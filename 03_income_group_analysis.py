"""
03_income_group_analysis.py

Pension Coverage & Healthy Life Expectancy (HALE) Study
Split-sample analysis by World Bank income group -- a direct test of H1
(that the pension-HALE relationship is stronger in lower income countries).

Runs both the contributory and social pension models separately for each
income group (Low, Lower middle, Upper middle, High), saves a formatted
Excel table comparing coefficients across groups, and flags groups where
the sample is too thin to produce reliable FE estimates.

INPUT:  cleaned_data/cleaned_panel.csv  (produced by 01_clean_data.py)
OUTPUT: outputs/income_group_analysis.xlsx

Run AFTER 01_clean_data.py and 02_analyze_data.py:
    python 03_income_group_analysis.py
"""

import pandas as pd
import numpy as np
from linearmodels.panel import PanelOLS
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
pd.set_option("display.width", 120)

# =============================================================================
# CONFIG
# =============================================================================

CLEANED_PANEL_PATH = "cleaned_data/cleaned_panel.csv"
OUT_DIR = "outputs"

DEPENDENT_VAR = "hale_60"
CONTROL_VARS = ["log_gdp_pc", "dependency_ratio", "oop_health_exp",
                 "gov_health_exp", "female_share_elderly", "urbanization_rate"]
ENTITY_COL = "iso3"
TIME_COL = "year"

# Minimum countries with 2+ years before we flag the estimate as unreliable
MIN_FE_COUNTRIES = 10

INCOME_GROUPS = ["Low income", "Lower middle income",
                 "Upper middle income", "High income"]

PENSION_MODELS = {
    "contributory": {
        "c_col":  "coverage_contrib_c",
        "sq_col": "coverage_contrib_c_sq",
        "label":  "Contributory",
    },
    "social": {
        "c_col":  "coverage_social_c",
        "sq_col": "coverage_social_c_sq",
        "label":  "Social (non-contributory)",
    },
}


# =============================================================================
# HELPERS
# =============================================================================

def load_panel():
    df = pd.read_csv(CLEANED_PANEL_PATH)
    df = df.set_index([ENTITY_COL, TIME_COL])
    return df


def stars(p):
    if pd.isna(p):   return ""
    if p < 0.01:     return "***"
    if p < 0.05:     return "**"
    if p < 0.10:     return "*"
    return ""


def run_group_model(df, group, model_spec):
    """
    Runs a two-way FE model for one income group and one pension type.
    Returns a dict of results, or a dict with a 'note' if sample is too thin.
    """
    c_col  = model_spec["c_col"]
    sq_col = model_spec["sq_col"]
    regressors = [c_col, sq_col] + CONTROL_VARS

    sub = df[df["income_group"] == group].dropna(subset=[DEPENDENT_VAR] + regressors)
    n_obs      = len(sub)
    n_countries = sub.index.get_level_values(ENTITY_COL).nunique()
    ypc        = sub.groupby(level=ENTITY_COL)[DEPENDENT_VAR].count()
    fe_countries = (ypc >= 2).sum()

    result = {
        "group":        group,
        "n_obs":        n_obs,
        "n_countries":  n_countries,
        "fe_countries": fe_countries,
        "coef":         np.nan,
        "se":           np.nan,
        "p":            np.nan,
        "sig":          "",
        "sq_coef":      np.nan,
        "sq_p":         np.nan,
        "r2":           np.nan,
        "note":         "",
    }

    if fe_countries < MIN_FE_COUNTRIES:
        result["note"] = f"Unreliable — only {fe_countries} countries with 2+ years"
        print(f"    {group}: SKIPPED ({fe_countries} FE-usable countries < {MIN_FE_COUNTRIES} threshold)")
        return result

    try:
        y   = sub[DEPENDENT_VAR]
        X   = sub[regressors]
        res = PanelOLS(y, X, entity_effects=True, time_effects=True,
                        drop_absorbed=True).fit(cov_type="clustered",
                                                 cluster_entity=True)
        result["coef"]   = round(res.params[c_col], 4)
        result["se"]     = round(res.std_errors[c_col], 4)
        result["p"]      = round(res.pvalues[c_col], 4)
        result["sig"]    = stars(res.pvalues[c_col])
        result["sq_coef"]= round(res.params[sq_col], 4)
        result["sq_p"]   = round(res.pvalues[sq_col], 4)
        result["r2"]     = round(res.rsquared, 4)
        print(f"    {group}: coef={result['coef']}, p={result['p']}, "
              f"N={n_obs}, countries={n_countries}")
    except Exception as e:
        result["note"] = f"Model failed: {e}"
        print(f"    {group}: FAILED — {e}")

    return result


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def write_excel(all_results, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Group Analysis"

    header_fill  = PatternFill("solid", start_color="1F4E5F", end_color="1F4E5F")
    subhead_fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
    warn_fill    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    alt_fill     = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    thin         = Side(style="thin")

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sf = Font(name="Arial", bold=True, size=10)
    bf = Font(name="Arial", bold=True, size=10)
    nf = Font(name="Arial", size=10)
    wf = Font(name="Arial", italic=True, size=9, color="7F6000")

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"].value = ("Split-Sample Analysis by Income Group — "
                      "Pension Coverage Effect on HALE at Age 60")
    ws["A1"].font  = Font(name="Arial", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Subtitle note
    ws.merge_cells("A2:I2")
    ws["A2"].value = ("Two-way fixed effects (country + year FE), SEs clustered at country level. "
                      "*** p<0.01  ** p<0.05  * p<0.10. "
                      f"Groups with <{MIN_FE_COUNTRIES} countries with 2+ years flagged as unreliable.")
    ws["A2"].font  = Font(name="Arial", italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    col_headers = ["Income Group", "Coverage Coef.", "Std. Error", "p-value", "Sig.",
                   "Coverage² Coef.", "Coverage² p", "R²", "N (obs) / Countries"]

    row = 4
    for model_key, model_spec in PENSION_MODELS.items():
        # Section header
        ws.merge_cells(f"A{row}:I{row}")
        sec = ws.cell(row=row, column=1,
                       value=f"Panel {model_key[0].upper()}: {model_spec['label']} Pension Coverage")
        sec.font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        sec.fill  = header_fill
        sec.alignment = Alignment(horizontal="left")
        row += 1

        # Column headers
        for c, h in enumerate(col_headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font      = hf
            cell.fill      = subhead_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border    = Border(bottom=thin)
        row += 1

        # Data rows
        for i, res in enumerate(all_results[model_key]):
            fill = alt_fill if i % 2 == 0 else None
            is_warn = bool(res["note"])

            vals = [
                res["group"],
                res["coef"]    if not is_warn else "—",
                res["se"]      if not is_warn else "—",
                res["p"]       if not is_warn else "—",
                res["sig"]     if not is_warn else "",
                res["sq_coef"] if not is_warn else "—",
                res["sq_p"]    if not is_warn else "—",
                res["r2"]      if not is_warn else "—",
                f"{res['n_obs']} / {res['n_countries']}" if not is_warn
                else res["note"],
            ]

            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font      = wf if is_warn else nf
                cell.alignment = Alignment(horizontal="left" if c == 1 else "center",
                                            wrap_text=(c == 9))
                if is_warn:
                    cell.fill = warn_fill
                elif fill:
                    cell.fill = fill
            row += 1

        row += 1  # blank row between panels

    # Column widths
    widths = [24, 14, 12, 10, 6, 16, 14, 8, 26]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A5"

    wb.save(out_path)
    print(f"\nSaved income group analysis table to {out_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    import os
    os.makedirs(OUT_DIR, exist_ok=True)

    df = load_panel()

    all_results = {}
    for model_key, model_spec in PENSION_MODELS.items():
        print(f"\n{'='*65}")
        print(f"MODEL: {model_spec['label'].upper()}")
        print(f"{'='*65}")
        group_results = []
        for group in INCOME_GROUPS:
            res = run_group_model(df, group, model_spec)
            group_results.append(res)
        all_results[model_key] = group_results

    out_path = f"{OUT_DIR}/income_group_analysis.xlsx"
    write_excel(all_results, out_path)

    # Print a clean comparison summary
    print("\n" + "="*65)
    print("SUMMARY — Coverage coefficient by income group")
    print("="*65)
    print(f"{'Group':<25} {'Contributory':>14} {'Social':>14}")
    print("-"*55)
    for i, group in enumerate(INCOME_GROUPS):
        c = all_results["contributory"][i]
        s = all_results["social"][i]
        c_str = f"{c['coef']}{c['sig']}" if not c["note"] else "n/a"
        s_str = f"{s['coef']}{s['sig']}" if not s["note"] else "n/a"
        print(f"{group:<25} {c_str:>14} {s_str:>14}")
    print("\n*** p<0.01  ** p<0.05  * p<0.10  |  n/a = sample too thin for FE")


if __name__ == "__main__":
    main()
